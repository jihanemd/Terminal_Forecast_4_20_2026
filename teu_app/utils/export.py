"""
export.py
---------
Génère un fichier Excel formaté à partir des résultats de prédiction batch.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date


# ── Couleurs ───────────────────────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="EEF4FF")
CORAL_FILL  = PatternFill("solid", fgColor="FDF2F2")
HEADER_FILL = PatternFill("solid", fgColor="111928")
WARN_FILL   = PatternFill("solid", fgColor="FFFBEB")
GRAY_FILL   = PatternFill("solid", fgColor="F9FAFB")

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)


def _header_style(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def _data_style(cell, fill=None):
    cell.font = Font(size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill


def export_to_excel(results: list) -> io.BytesIO:
    """
    Crée un fichier Excel formaté avec les résultats de prédiction.
    Retourne un buffer BytesIO prêt à être envoyé.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Prédictions L-D"

    # ── Titre ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"Terminal L/D Predictor — Export du {date.today().strftime('%d/%m/%Y')}"
    title_cell.font  = Font(bold=True, size=13, color="111928", name="Calibri")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Sous-titre ─────────────────────────────────────────────────────────────
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = "Ratios basés sur les 3 derniers mois d'activité par lane"
    sub.font  = Font(size=10, color="6B7280", name="Calibri")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8  # Espace

    # ── En-têtes ───────────────────────────────────────────────────────────────
    headers = [
        "Lane", "Volume navire",
        "Load estimé", "% Load",
        "Discharge estimé", "% Discharge",
        "Période de référence", "N escales", "Statut"
    ]
    for col, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=col), h)
    ws.row_dimensions[4].height = 22

    # ── Données ────────────────────────────────────────────────────────────────
    for row_idx, res in enumerate(results, start=5):
        unknown  = res.get("unknown", False)
        outdated = res.get("outdated", False)

        # Lane
        c = ws.cell(row=row_idx, column=1, value=res["lane"])
        c.font = Font(bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        # Volume
        c = ws.cell(row=row_idx, column=2, value=res["volume"])
        _data_style(c)
        c.number_format = "#,##0"

        if unknown:
            # Lane inconnue → cellules grises
            for col in range(3, 10):
                c = ws.cell(row=row_idx, column=col, value="—" if col != 9 else "Lane inconnue")
                _data_style(c, GRAY_FILL)
        else:
            # Load
            c = ws.cell(row=row_idx, column=3, value=res["pred_L"])
            _data_style(c, BLUE_FILL)
            c.number_format = "#,##0"
            c.font = Font(bold=True, size=10, color="1E40AF", name="Calibri")

            # % Load
            c = ws.cell(row=row_idx, column=4, value=res["pct_L"] / 100)
            _data_style(c, BLUE_FILL)
            c.number_format = "0.0%"
            c.font = Font(size=10, color="1E40AF", name="Calibri")

            # Discharge
            c = ws.cell(row=row_idx, column=5, value=res["pred_D"])
            _data_style(c, CORAL_FILL)
            c.number_format = "#,##0"
            c.font = Font(bold=True, size=10, color="991B1B", name="Calibri")

            # % Discharge
            c = ws.cell(row=row_idx, column=6, value=res["pct_D"] / 100)
            _data_style(c, CORAL_FILL)
            c.number_format = "0.0%"
            c.font = Font(size=10, color="991B1B", name="Calibri")

            # Période
            period = f"{res['period_start']} → {res['last_date']}"
            c = ws.cell(row=row_idx, column=7, value=period)
            _data_style(c)
            c.alignment = Alignment(horizontal="left", vertical="center")

            # N escales
            c = ws.cell(row=row_idx, column=8, value=res["n_voyages"])
            _data_style(c)

            # Statut
            if outdated:
                c = ws.cell(row=row_idx, column=9, value="⚠ Données anciennes")
                _data_style(c, WARN_FILL)
                c.font = Font(size=10, color="92400E", name="Calibri")
            else:
                c = ws.cell(row=row_idx, column=9, value="OK")
                _data_style(c)
                c.font = Font(size=10, color="057A55", name="Calibri")

        ws.row_dimensions[row_idx].height = 20

    # ── Largeurs des colonnes ──────────────────────────────────────────────────
    col_widths = [12, 16, 16, 10, 18, 13, 30, 12, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes (figer l'en-tête) ────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Sauvegarde dans buffer ─────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ── Couleurs F/M (ajoutées pour la fonctionnalité Full/Empty) ─────────────────
GREEN_FILL = PatternFill("solid", fgColor="ECFDF5")
AMBER_FILL = PatternFill("solid", fgColor="FFFBEB")


def export_fm_to_excel(results: list) -> io.BytesIO:
    """
    Exporte les résultats du pipeline complet (D/L + F/M) en Excel formaté.
    Colonnes : Lane, Volume, Discharge, Load, D-Full, D-Empty, L-Full, L-Empty.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Prédictions L-D + F-M"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value = f"Terminal Predictor — L/D + Full/Empty — Export du {date.today().strftime('%d/%m/%Y')}"
    t.font  = Font(bold=True, size=13, color="111928", name="Calibri")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    s = ws["A2"]
    s.value = "Pipeline : Volume → Load/Discharge (ratios D/L) → Full/Empty (ratios F/M, 3 derniers mois)"
    s.font  = Font(size=10, color="6B7280", name="Calibri")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    headers = ["Lane", "Volume", "Discharge", "Load",
               "D-Full", "% D-Full", "D-Empty", "% D-Empty",
               "L-Full", "% L-Full", "L-Empty", "% L-Empty"]
    for col, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=4, column=col), h)
    ws.row_dimensions[4].height = 22

    for row_idx, res in enumerate(results, start=5):
        unknown  = res.get("unknown", False)
        outdated = res.get("outdated", False)
        fm_ok    = res.get("fm_available", False)

        c = ws.cell(row=row_idx, column=1, value=res["lane"])
        c.font = Font(bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=row_idx, column=2, value=res["volume"])
        _data_style(c); c.number_format = "#,##0"

        if unknown:
            for col in range(3, 13):
                c = ws.cell(row=row_idx, column=col, value="Lane inconnue" if col == 3 else "—")
                _data_style(c, GRAY_FILL)
        else:
            # Discharge total
            c = ws.cell(row=row_idx, column=3, value=res.get("pred_D"))
            _data_style(c, CORAL_FILL); c.number_format = "#,##0"
            c.font = Font(bold=True, size=10, color="991B1B", name="Calibri")

            # Load total
            c = ws.cell(row=row_idx, column=4, value=res.get("pred_L"))
            _data_style(c, BLUE_FILL); c.number_format = "#,##0"
            c.font = Font(bold=True, size=10, color="1E40AF", name="Calibri")

            if not fm_ok:
                for col in range(5, 13):
                    c = ws.cell(row=row_idx, column=col, value="N/D")
                    _data_style(c, GRAY_FILL)
            else:
                data_fm = [
                    (res.get("D_full"),  res.get("pct_D_full"),  CORAL_FILL, "991B1B"),
                    (res.get("D_empty"), res.get("pct_D_empty"), AMBER_FILL, "92400E"),
                    (res.get("L_full"),  res.get("pct_L_full"),  BLUE_FILL,  "1E40AF"),
                    (res.get("L_empty"), res.get("pct_L_empty"), GREEN_FILL, "065F46"),
                ]
                col = 5
                for val, pct, fill, color in data_fm:
                    c = ws.cell(row=row_idx, column=col, value=val)
                    _data_style(c, fill); c.number_format = "#,##0"
                    c.font = Font(bold=True, size=10, color=color, name="Calibri")
                    col += 1
                    c = ws.cell(row=row_idx, column=col, value=(pct or 0) / 100)
                    _data_style(c, fill); c.number_format = "0.0%"
                    c.font = Font(size=10, color=color, name="Calibri")
                    col += 1

        ws.row_dimensions[row_idx].height = 20

    for i, w in enumerate([10, 10, 12, 12, 10, 9, 10, 9, 10, 9, 10, 9], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_teu_to_excel(results: list) -> io.BytesIO:
    """
    Exporte les résultats du pipeline complet (D/L + F/M + TEU) en Excel.
    Colonnes : Lane, Volume, D, L, D-Full, D-Empty, L-Full, L-Empty,
               D-Full TEU, D-Empty TEU, L-Full TEU, L-Empty TEU, Total TEU.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Prédictions TEU Complet"

    # Couleurs TEU
    teu_blue_fill  = PatternFill("solid", fgColor="DBEAFE")   # L-Full TEU
    teu_green_fill = PatternFill("solid", fgColor="D1FAE5")   # L-Empty TEU
    teu_red_fill   = PatternFill("solid", fgColor="FFE4E6")   # D-Full TEU
    teu_amber_fill = PatternFill("solid", fgColor="FEF3C7")   # D-Empty TEU
    total_fill     = PatternFill("solid", fgColor="EDE9FE")   # Total TEU

    headers = [
        "Lane", "Volume",
        "Discharge", "Load",
        "D-Full", "D-Empty", "L-Full", "L-Empty",
        "D-Full TEU", "D-Empty TEU", "L-Full TEU", "L-Empty TEU",
        "Total TEU"
    ]
    WHITE_HEADER_FILL = PatternFill("solid", fgColor="FFFFFF")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True, size=10, color="111928", name="Calibri")
        cell.fill = WHITE_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    for row_idx, res in enumerate(results, start=2):
        unknown  = res.get("unknown", False)
        outdated = res.get("outdated", False)
        fm_ok    = res.get("fm_available", False)
        teu_ok   = res.get("teu_available", False)

        c = ws.cell(row=row_idx, column=1, value=res.get("lane", "?"))
        c.font = Font(bold=True, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        c = ws.cell(row=row_idx, column=2, value=res.get("volume"))
        _data_style(c); c.number_format = "#,##0"

        if unknown:
            for col in range(3, 14):
                c = ws.cell(row=row_idx, column=col, value="Lane inconnue" if col == 3 else "—")
                _data_style(c, GRAY_FILL)
        else:
            # D et L containers total
            c = ws.cell(row=row_idx, column=3, value=res.get("pred_D"))
            _data_style(c, CORAL_FILL); c.number_format = "#,##0"
            c.font = Font(bold=True, size=10, color="991B1B", name="Calibri")

            c = ws.cell(row=row_idx, column=4, value=res.get("pred_L"))
            _data_style(c, BLUE_FILL); c.number_format = "#,##0"
            c.font = Font(bold=True, size=10, color="1E40AF", name="Calibri")

            # F/M containers
            fm_vals = [
                (res.get("D_full"),  CORAL_FILL, "991B1B"),
                (res.get("D_empty"), WARN_FILL,  "92400E"),
                (res.get("L_full"),  BLUE_FILL,  "1E40AF"),
                (res.get("L_empty"), GREEN_FILL, "065F46"),
            ]
            for i, (val, fill, color) in enumerate(fm_vals, start=5):
                c = ws.cell(row=row_idx, column=i, value=val if fm_ok else "N/D")
                _data_style(c, fill if fm_ok else GRAY_FILL)
                if fm_ok and val is not None:
                    c.number_format = "#,##0"
                    c.font = Font(bold=True, size=10, color=color, name="Calibri")

            # TEU par groupe
            teu_vals = [
                (res.get("D_full_teu"),  teu_red_fill,   "7F1D1D"),
                (res.get("D_empty_teu"), teu_amber_fill, "78350F"),
                (res.get("L_full_teu"),  teu_blue_fill,  "1E3A8A"),
                (res.get("L_empty_teu"), teu_green_fill, "065F46"),
            ]
            for i, (val, fill, color) in enumerate(teu_vals, start=9):
                c = ws.cell(row=row_idx, column=i, value=val if teu_ok else "N/D")
                _data_style(c, fill if teu_ok else GRAY_FILL)
                if teu_ok and val is not None:
                    c.number_format = "#,##0"
                    c.font = Font(bold=True, size=10, color=color, name="Calibri")

            # Total TEU
            c = ws.cell(row=row_idx, column=13, value=res.get("total_teu") if teu_ok else "N/D")
            _data_style(c, total_fill if teu_ok else GRAY_FILL)
            if teu_ok and res.get("total_teu") is not None:
                c.number_format = "#,##0"
                c.font = Font(bold=True, size=11, color="4C1D95", name="Calibri")

        ws.row_dimensions[row_idx].height = 20

    for i, w in enumerate([10, 10, 12, 12, 10, 10, 10, 10, 12, 13, 12, 13, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_teu_summary_to_excel(result: dict) -> io.BytesIO:
    """
    Export TEU synthèse (lane unique) dans le format hiérarchique Excel :
      L1 : Service | Date | Forecast Volume | Discharge (×3) | Load (×3)
      L2 :                                  | Total D         | Total L
      L3 :                                  | Full|Empty|Reefer | Full|Empty|Reefer
      L4 :                                  | containers counts
      L5 :                                  | TEU | TEU | TEU | TEU | TEU | TEU
      L6 : BSMAR | date | volume            | TEU values
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse TEU"

    DARK_NAVY  = PatternFill("solid", fgColor="132954")
    MED_BLUE   = PatternFill("solid", fgColor="2563EB")
    LIGHT_BLUE = PatternFill("solid", fgColor="DBEAFE")
    TOTAL_FILL = PatternFill("solid", fgColor="EEF4FF")
    LEFT_FILL  = PatternFill("solid", fgColor="EBF0FA")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    DATA_FILL  = PatternFill("solid", fgColor="F0F7FF")

    def _navy(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
        cell.fill = DARK_NAVY
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _med_blue(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill = MED_BLUE
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _total(cell, val):
        cell.value = val if val is not None else 0
        cell.font = Font(bold=True, color="1A56DB", size=16, name="Calibri")
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        cell.number_format = "#,##0"

    def _cntr(cell, val):
        if val is None:
            cell.value = "-"
            cell.font = Font(color="9CA3AF", size=12, name="Calibri")
        else:
            cell.value = val
            cell.font = Font(bold=True, size=13, name="Calibri")
            cell.number_format = "#,##0"
        cell.fill = WHITE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _teu_lbl(cell):
        cell.value = "TEU"
        cell.font = Font(bold=True, color="1D4ED8", size=10, name="Calibri")
        cell.fill = LIGHT_BLUE
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _teu_val(cell, val):
        if val is None:
            cell.value = "-"
            cell.font = Font(color="9CA3AF", size=12, name="Calibri")
        else:
            cell.value = val
            cell.font = Font(bold=True, size=13, name="Calibri")
            cell.number_format = "#,##0"
        cell.fill = DATA_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    def _left_label(cell, text):
        cell.value = text
        cell.font = Font(bold=True, size=11, color="374151", name="Calibri")
        cell.fill = LEFT_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    def _left_data(cell, val, bold=False, fmt=None):
        cell.value = val
        cell.font = Font(bold=bold, size=11, name="Calibri")
        cell.fill = WHITE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        if fmt:
            cell.number_format = fmt

    fm_ok  = result.get("fm_available",  False)
    teu_ok = result.get("teu_available", False)

    # Colonnes A-C : labels (lignes 1-5 fusionnées)
    ws.merge_cells("A1:A5"); _left_label(ws["A1"], "Service")
    ws.merge_cells("B1:B5"); _left_label(ws["B1"], "Date")
    ws.merge_cells("C1:C5"); _left_label(ws["C1"], "Forecast\nvolume")

    # Ligne 1 — Discharge / Load
    ws.merge_cells("D1:F1"); _navy(ws["D1"], "Discharge")
    ws.merge_cells("G1:I1"); _navy(ws["G1"], "Load")

    # Ligne 2 — Totaux containers
    ws.merge_cells("D2:F2"); _total(ws["D2"], result.get("pred_D"))
    ws.merge_cells("G2:I2"); _total(ws["G2"], result.get("pred_L"))

    # Ligne 3 — Full | Empty | Reefer
    for col, lbl in zip([4,5,6,7,8,9], ["Full","Empty","Reefer","Full","Empty","Reefer"]):
        _med_blue(ws.cell(row=3, column=col), lbl)

    # Ligne 4 — Containers
    d_full  = result.get("D_full")  if fm_ok else None
    d_empty = result.get("D_empty") if fm_ok else None
    l_full  = result.get("L_full")  if fm_ok else None
    l_empty = result.get("L_empty") if fm_ok else None
    for col, val in zip([4,5,6,7,8,9], [d_full, d_empty, None, l_full, l_empty, None]):
        _cntr(ws.cell(row=4, column=col), val)

    # Ligne 5 — Labels TEU
    for col in range(4, 10):
        _teu_lbl(ws.cell(row=5, column=col))

    # Ligne 6 — Données réelles
    _left_data(ws.cell(row=6, column=1), result.get("lane", ""), bold=True)
    _left_data(ws.cell(row=6, column=2), result.get("last_date", ""))
    _left_data(ws.cell(row=6, column=3), result.get("volume"), bold=True, fmt="#,##0")

    d_ft = result.get("D_full_teu")  if teu_ok else None
    d_et = result.get("D_empty_teu") if teu_ok else None
    l_ft = result.get("L_full_teu")  if teu_ok else None
    l_et = result.get("L_empty_teu") if teu_ok else None
    for col, val in zip([4,5,6,7,8,9], [d_ft, d_et, None, l_ft, l_et, None]):
        _teu_val(ws.cell(row=6, column=col), val)

    # Dimensions
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for col in ["D","E","F","G","H","I"]:
        ws.column_dimensions[col].width = 13

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 26

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
